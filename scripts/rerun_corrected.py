#!/usr/bin/env python3
"""Re-run image sourcing for user-flagged SKUs in CORRECTED.xlsx.

Orange highlight (FFFFC000) = wrong wine photo — strict label match + quality.
Yellow highlight (FFFFFF00) = right wine but unusable — strict quality (watermarks).
Also fills any missing images from S004725 onward.
"""

from __future__ import annotations

import argparse
import asyncio
import logging
import shutil
import sys
import zipfile
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

_REPO = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(_REPO / "src"))
sys.path.insert(0, str(_REPO / "scripts"))

from dotenv import load_dotenv

load_dotenv()

from process_pricelist import (  # noqa: E402
    _base_wine_key,
    _extract_producer,
    _publish_image,
    _safe_sku_filename,
)
from wine_pipeline.cache import ResultCache
from wine_pipeline.label_extractor import LabelExtractor
from wine_pipeline.models import SKU
from wine_pipeline.ocr import OCRCrossChecker
from wine_pipeline.pipeline import Pipeline
from wine_pipeline.quality_filter import QualityFilter
from wine_pipeline.scoring import ConfidenceScorer
from wine_pipeline.search import SearchModule
from wine_pipeline.verifier import FingerprintVerifier

logger = logging.getLogger("rerun_corrected")

ORANGE_RGB = "FFFFC000"
YELLOW_RGB = "FFFFFF00"


def _cell_rgb(cell) -> str | None:
    fill = cell.fill
    if fill is None or fill.fill_type != "solid":
        return None
    color = fill.start_color
    if color and color.type == "rgb" and color.rgb:
        return color.rgb.upper()
    return None


def parse_highlighted_skus(excel_path: Path) -> tuple[list[str], list[str]]:
    """Return (orange_sku_ids, yellow_sku_ids) from fill colors."""
    wb = load_workbook(excel_path)
    ws = wb.active
    orange: list[str] = []
    yellow: list[str] = []
    for row in range(2, ws.max_row + 1):
        code = ws.cell(row, 1).value
        if not code:
            continue
        sku_id = str(code).strip()
        colors = {
            c
            for col in range(1, ws.max_column + 1)
            if (c := _cell_rgb(ws.cell(row, col)))
        }
        if ORANGE_RGB in colors:
            orange.append(sku_id)
        elif YELLOW_RGB in colors:
            yellow.append(sku_id)
    return orange, yellow


def load_skus_from_excel(excel_path: Path, sku_ids: set[str]) -> list[SKU]:
    df = pd.read_excel(excel_path)
    code_col = "wine_id" if "wine_id" in df.columns else "Code"
    name_col = "full_wine_name" if "full_wine_name" in df.columns else "Name"
    vintage_col = "vintage" if "vintage" in df.columns else "Vintage"
    volume_col = next((c for c in ("format", "Volume") if c in df.columns), None)
    type_col = next((c for c in ("wine_type", "Type") if c in df.columns), None)
    country_col = next((c for c in ("country", "Country") if c in df.columns), None)
    grape_col = next((c for c in ("grapes", "Grape") if c in df.columns), None)
    region_col = next(
        (c for c in ("sub_region", "region", "Region") if c in df.columns),
        None,
    )

    skus: list[SKU] = []
    for _, row in df.iterrows():
        code = str(row[code_col])
        if code not in sku_ids:
            continue
        name = str(row[name_col])
        vintage_str = str(row[vintage_col]) if pd.notna(row[vintage_col]) else None
        vintage = None
        if vintage_str and vintage_str.upper() != "NV":
            try:
                vintage = int(float(vintage_str))
            except ValueError:
                pass
        region = ""
        if region_col and pd.notna(row.get(region_col)):
            region = str(row[region_col]).strip()
        grape = ""
        if grape_col and pd.notna(row.get(grape_col)):
            grape = str(row[grape_col])
        skus.append(
            SKU(
                id=code,
                producer=_extract_producer(name),
                appellation=region,
                cru_vineyard=name,
                vintage=vintage,
                format=str(row[volume_col])
                if (volume_col and pd.notna(row.get(volume_col)))
                else "750ml",
                region=region,
                full_name=name,
                wine_type=str(row[type_col])
                if (type_col and pd.notna(row.get(type_col)))
                else None,
                country=str(row[country_col])
                if (country_col and pd.notna(row.get(country_col)))
                else None,
                grapes=grape,
            )
        )
    return skus


def _remove_image(sku_id: str, *folders: Path) -> None:
    safe = _safe_sku_filename(sku_id)
    for folder in folders:
        path = folder / f"{safe}.jpg"
        if path.is_file():
            path.unlink()


def _configure_pipeline(pipeline: Pipeline, mode: str) -> None:
    pipeline._skip_ocr = True
    pipeline._skip_removebg = True
    pipeline._skip_quality = False
    pipeline.search._vivino_first_only = True

    if mode == "orange":
        pipeline._strict_match = True
        pipeline._min_producer_match = 0.50
        pipeline._min_vintage_match = 0.75
        pipeline._strict_quality = True
        pipeline._require_pass_verdict = True
    elif mode == "yellow":
        pipeline._strict_match = False
        pipeline._strict_quality = True
        pipeline._require_pass_verdict = True
    else:
        pipeline._strict_match = False
        pipeline._strict_quality = False
        pipeline._require_pass_verdict = False


async def _process_one(
    pipeline: Pipeline,
    sku: SKU,
    img_cache: Path,
    out_dir: Path,
    base_wine_images: dict[str, Path],
    max_candidates: int,
    sku_timeout_sec: int,
    *,
    allow_reuse: bool = True,
) -> bool:
    safe_id = _safe_sku_filename(sku.id)
    src_path = img_cache / f"{safe_id}.jpg"
    base_key = _base_wine_key(sku.full_name or "")

    if allow_reuse and base_key and base_key in base_wine_images:
        shutil.copy2(base_wine_images[base_key], src_path)
        _publish_image(sku.id, src_path, out_dir)
        logger.info("  ↻ Reused image from same wine (different vintage)")
        return True

    _remove_image(sku.id, img_cache, out_dir)

    try:
        result = await asyncio.wait_for(
            pipeline.process_sku(
                sku, bypass_cache=True, max_candidates=max_candidates
            ),
            timeout=sku_timeout_sec,
        )
    except asyncio.TimeoutError:
        logger.error("  Timeout (%ds) for %s", sku_timeout_sec, sku.id)
        return False
    except Exception as exc:
        logger.error("  Error processing %s: %s", sku.id, exc)
        return False

    if src_path.is_file() and src_path.stat().st_size > 0:
        if base_key:
            base_wine_images[base_key] = src_path
        _publish_image(sku.id, src_path, out_dir)
        logger.info(
            "  ✓ %s confidence=%.0f%% producer=%.0f%% vintage=%.0f%%",
            sku.id,
            result.overall_confidence * 100,
            result.producer_match * 100,
            result.vintage_match * 100,
        )
        return True

    logger.warning(
        "  No image for %s (verdict=%s, conf=%.0f%%)",
        sku.id,
        result.verdict.value,
        result.overall_confidence * 100,
    )
    return False


async def rerun_corrected(
    excel_path: str,
    output_dir: str,
    zip_name: str,
    max_candidates: int = 15,
    sku_timeout_sec: int = 120,
    from_sku: str = "S004725",
) -> None:
    excel = Path(excel_path)
    out_dir = Path(output_dir)
    img_cache = Path("data/images")
    out_dir.mkdir(parents=True, exist_ok=True)
    img_cache.mkdir(parents=True, exist_ok=True)

    orange_ids, yellow_ids = parse_highlighted_skus(excel)
    logger.info(
        "CORRECTED highlights: %d orange (wrong), %d yellow (low quality)",
        len(orange_ids),
        len(yellow_ids),
    )

    df = pd.read_excel(excel)
    code_col = "wine_id" if "wine_id" in df.columns else "Code"
    all_codes = [str(c) for c in df[code_col]]
    missing_after = [
        c
        for c in all_codes
        if c > from_sku
        and not (out_dir / f"{_safe_sku_filename(c)}.jpg").is_file()
    ]
    logger.info(
        "Missing after %s: %d SKUs (%s)",
        from_sku,
        len(missing_after),
        ", ".join(missing_after[:8]) + ("..." if len(missing_after) > 8 else ""),
    )

    orange_set = set(orange_ids)
    yellow_set = set(yellow_ids)
    missing_set = set(missing_after) - orange_set - yellow_set

    pipeline = Pipeline(
        search=SearchModule(),
        label_extractor=LabelExtractor(),
        verifier=FingerprintVerifier(),
        ocr_checker=OCRCrossChecker(),
        quality_filter=QualityFilter(),
        scorer=ConfidenceScorer(),
        cache=ResultCache(":memory:"),
    )

    base_wine_images: dict[str, Path] = {}
    for folder in (img_cache, out_dir):
        for path in folder.glob("*.jpg"):
            if path.stat().st_size > 0:
                base_wine_images.setdefault(path.stem, path)

    groups = [
        ("orange", orange_ids),
        ("yellow", yellow_ids),
        ("missing", sorted(missing_set)),
    ]

    stats = {"orange": 0, "yellow": 0, "missing": 0}
    for mode, sku_id_list in groups:
        if not sku_id_list:
            continue
        skus = load_skus_from_excel(excel, set(sku_id_list))
        _configure_pipeline(pipeline, mode)
        logger.info("=== %s: re-running %d SKUs ===", mode.upper(), len(skus))
        for idx, sku in enumerate(skus):
            logger.info("[%d/%d] %s: %s", idx + 1, len(skus), sku.id, sku.full_name)
            if await _process_one(
                pipeline,
                sku,
                img_cache,
                out_dir,
                base_wine_images,
                max_candidates,
                sku_timeout_sec,
                allow_reuse=(mode != "orange"),
            ):
                stats[mode] += 1

    logger.info("Copying all images and rebuilding ZIP...")
    count = 0
    for path in sorted(out_dir.glob("*.jpg")):
        if path.stat().st_size > 0:
            count += 1

    with zipfile.ZipFile(zip_name, "w", zipfile.ZIP_DEFLATED) as zf:
        for img in sorted(out_dir.glob("*.jpg")):
            if img.stat().st_size > 0:
                zf.write(img, img.name)

    logger.info(
        "Done — orange %d/%d, yellow %d/%d, missing %d/%d, %d total in ZIP",
        stats["orange"],
        len(orange_ids),
        stats["yellow"],
        len(yellow_ids),
        stats["missing"],
        len(missing_set),
        count,
    )


def main() -> None:
    logging.basicConfig(
        level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s"
    )
    parser = argparse.ArgumentParser(description="Re-run CORRECTED.xlsx flagged SKUs")
    parser.add_argument("--excel", default="CORRECTED.xlsx")
    parser.add_argument("--output-dir", default="artifacts/Goldgate_wine")
    parser.add_argument("--zip", default="artifacts/Goldgate_wine.zip")
    parser.add_argument("--from-sku", default="S004725")
    parser.add_argument("--max-candidates", type=int, default=15)
    parser.add_argument("--timeout", type=int, default=120)
    args = parser.parse_args()
    asyncio.run(
        rerun_corrected(
            args.excel,
            args.output_dir,
            args.zip,
            max_candidates=args.max_candidates,
            sku_timeout_sec=args.timeout,
            from_sku=args.from_sku,
        )
    )


if __name__ == "__main__":
    main()
